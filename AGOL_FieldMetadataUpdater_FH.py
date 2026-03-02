# Script name: ArcGIS Online Field Information Updater
# Created by: Farah Hoque, Esri Canada (2026) inspired by Lisa Berry's Hosted Feature Service Alias Updater script: https://github.com/lisaberrygis/AliasUpdater/tree/main

# Who is this script for?
# Anyone who wants to make their feature services more user-friendly and AI compatible.
# This script will also help recover a feature service's information (alias/description/value type) when it is overwritten and its field information gets wiped out

# What can this script do?
# Creates a lookup table for the field alias, description and value type of fields in a service which you can update and publish the changes on AGOL

# This script has 2 main functions:
# Extracts existing alias, description, and value type of fields in a feature service containing one or more layers and saves as a lookup table inside a folder. For a feature service with multiple layers, the lookup table will save each layer as a sheet with the layer name and layer ID.
# Updates the service on ArcGIS Online using the lookup table.

# Requirements:
# 1) You must have ArcGIS Pro installed on your computer
# 2) You must own the service you are trying to update

#------------ Instructions: ------------------

# 1) Go to PART 3 of the main script and enter all the parameters. Once PART 1 of the script executes, you will get prompted to review the lookup table to continue.
# If you are trying to make modifications to an existing lookup table, comment out PART 1 of the main script and re-run


# Import modules
from arcgis.gis import GIS
import pandas as pd
import openpyxl
import os
import ast
import re

#################################### PART 1: CONSTRUCT THE LOOKUP TABLE USING EXISTING FIELDS IN THE SERVICE ##############

def createLookupTable(gis, itemID, lookupTablepath, file_name):

    # This function converts the Esri recognized field type to the JSON backend value type
    def getValuetype(esri_type):
        # Map Esri field type to JSON type

        # String types
        if esri_type in ("esriFieldTypeString", "esriFieldTypeXML"):
            backend_types = [
                "nameOrTitle", "description", "typeOrCategory",
                "locationOrPlaceName", "phoneNumber", "emailAddress", "uniqueIdentifier", "dateAndTime", "coordinate", "binary"
            ]
            return 'Choose a value type from this list: ' + ', '.join(backend_types)

        # Integer types
        elif esri_type in ("esriFieldTypeBigInteger", "esriFieldTypeInteger" or "esriFieldTypeSmallInteger"):
            backend_types = ["countOrAmount", "orderedOrRanked", "binary", "uniqueIdentifier"]
            return 'Choose a value type from this list: ' + ', '.join(backend_types)

        # Float types
        elif esri_type in ("esriFieldTypeDouble", "esriFieldTypeSingle"):
            backend_types = ["percentageOrRatio", "measurement", "currency", "coordinate", "countOrAmount", "uniqueIdentifier"]
            return 'Choose a value type from this list: ' + ', '.join(backend_types)

        # Unique ID types
        elif esri_type in ("esriFieldTypeGUID", "esriFieldTypeOID", "esriFieldTypeGlobalID"):
            backend_types = ["uniqueIdentifier"]
            return backend_types[0]

        # Date and time types
        elif esri_type in ("esriFieldTypeDate", "esriFieldTypeDateOnly", "esriFieldTypeTimeOnly", "esriFieldTypeTimestampOffset"):
            backend_types = ["dateAndTime"]
            return backend_types[0]

        else:
            backend_types = ''  # Unknown or unsupported type
            return backend_types

    # Grab the service from the item ID, loop through each field
    item = gis.content.get(itemID)

    # Prepare an Excel writer. The field information needs to be saved to an Excel file rather than csv so it can
    # create multiple sheets for multiple layers
    writer = pd.ExcelWriter(lookupTablepath+file_name, engine="openpyxl")


    # Loop through the layers and grab existing field alias, description and value types if available
    for layer in item.layers:
        layer_name = layer.properties.name
        print(">>> Creating lookup table for '{}'".format(layer_name))

        # Define the sheet name
        # The layer_name is truncated to 28 characters because Excel only allows 31 characters for the sheet name.
        # The layer id is appended at the end of the sheet name
        sheet_name = layer_name[:28] + '_' + str(layer.properties.id)


        # Build the rows containing information for each field
        rows = []
        for field in layer.properties.fields:
            # Checks whether the field has a description. If so, populate it in the xlsx otherwise leave blank
            description = ast.literal_eval(desc).get("value", "") if isinstance((desc := dict(field).get("description")), str) and desc.strip() else ""

            # This checks if the description field already has value types set.
            # If it does, it grabs that value, otherwise gives a list of options from the getValuetype function
            valueTypeEnter = getValuetype(field.get("type"))
            getValueType = ast.literal_eval(desc).get("fieldValueType", "") if isinstance((desc := dict(field).get("description")), str) and desc.strip() else valueTypeEnter

            # Write the field information to the rows list
            rows.append({
            "Field": field.get("name"),
            "Alias": field.get("alias") if field.get("alias") else "",
            "Description": description,
            "ValueType": '' if field.get("name") in ('Shape__Area', 'Shape__Length', 'SHAPE__Area', 'SHAPE__Length', 'OBJECTID', 'FID') else getValueType
            })

        print('>>> Extracting field information: {}'.format(rows))

        # Populate an empty DataFrame with the layer's field informaton in the rows list
        df = pd.DataFrame(rows, columns=['Field', 'Alias', 'Description', 'ValueType'])

        # Save the dataframe under the appropriate sheet name
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f">>> Done writing {len(df)} fields to sheet: {sheet_name}\n\n")

    # Close the workbook
    writer.close()
    print(f"\nSaved the Excel file to path: {lookupTablepath+file_name}")


################################## PART 2: UPDATE THE LAYER'S INFORMATION USING THE LOOKUP TABLE #####################################

def updateitemAGOL(lookupTablepath, file_name, itemID, gis):

    # Join the lookup table folder path and file name to retrieve the file
    lookupTable = os.path.join(lookupTablepath, file_name)

    # Load the Excel workbook
    workbook = openpyxl.load_workbook(lookupTable)

    # Loop through ALL the sheets in the workbook that correspond to each layer in the feature service
    for sheet_name in workbook.sheetnames:

        sheet = workbook[sheet_name]
        print(f"\n\n-------> Processing sheet: {sheet_name}\n")

        # Create a list to store each field's alias, description and value type provided in each row. min_row=2 skips the first header row
        sheet_layer_fields = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True)]

        # Access the feature service to update
        item = gis.content.get(itemID)


        # Loop through each layer in the service, grab the layer id and match it to the sheet name
        for counter, layer in enumerate(item.layers):
            layer_id = layer.properties.id

            # Obtain the layer number from the sheet name by using regex
            sheet_layer_id = int(re.search(r"(\d+)$", sheet_name).group(1))

            # Match layer to the corresponding sheet by the id number
            if layer_id == sheet_layer_id:
                print(">>> Updating layer {}: ".format(str(counter+1)) + str(layer.properties.name))

                AGOL_layer_fields = layer.manager.properties.fields

                # Loop through the fields in the service and store in a list
                updateItemJSON = []

                for AGOL_field in AGOL_layer_fields:
                    field_name = AGOL_field['name']

                    # Match each field in the layer to the corresponding field in the sheet
                    for sheet_field in sheet_layer_fields:
                        if sheet_field[0] == field_name:
                            # If there is a match, store the field's json in a dict called fieldJSON
                            fieldJSON = dict(AGOL_field)

                            # Overwrite the new alias, description, and value types to the fieldJSON dict
                            alias = sheet_field[1]
                            fieldJSON['alias'] = alias

                            # If values exist for the description and value types, then populate otherwise leave blank
                            description_original = sheet_field[2] if sheet_field[2] else ""

                            # Modifies description to exclude special characters, new lines, tabs, and quotations
                            if description_original != "":
                                description_original = description_original.replace('\\n', ' ').replace('\\t', ' ')
                                # Removes everything except letters, numbers, commas, spaces and periods.
                                description = re.sub(r'[^A-Za-z0-9 ,.]+', '', description_original)
                                # Collapse multiple spaces
                                description = re.sub(r'\s+', ' ', description).strip()
                            else:
                                description = description_original

                            value_type = sheet_field[3] if sheet_field[3] else ""

                            fieldJSON['description'] = "{\"value\":" + "\"" + description + "\"" + ",\"fieldValueType\":\"" + value_type + "\"}"

                            print("\t The field {} had the following updates: ".format(field_name))
                            print("\t\t Alias: {}".format(alias))
                            print("\t\t Description: {}".format(description))
                            print("\t\t Value Type: {}".format(value_type))

                            # Append the fieldJSON dict to the updateItemJSON list
                            updateItemJSON.append(fieldJSON)


                fieldInfoUpdateDict = {'fields': updateItemJSON}

                # Use the update definition call to push the new field information to the service
                layer.manager.update_definition(fieldInfoUpdateDict)
                print("The alias, description and value types were updated on the layer '{}'!".format(layer.properties.name))

    print('\n>>> Finished updating all the layers in this service on ArcGIS Online!')


#################################### PART 3: DEFINE THE 4 PARAMETERS AND RUN MAIN SCRIPT #####################################

def main():
    ### 1. Define the username, password, and optionally the portal name
    gis = GIS(username="", password="")
    ### 2. Provide the item id
    itemID = ""
    ### 3. Provide the path to the folder that will store the lookup table(s). Example: "C:\Users\[username]\Desktop\\". Please include a double-slash at the end
    lookupTablepath = r""
    ### 4. Provide a file name for the lookup table. Make sure to include at the end .xlsx
    file_name = ""

    # Part 1: Create lookup table.
    # If you are making modifications to an existing lookup table and pushing the changes to AGOL, comment this part 1 portion out and re-run the script

    createLookupTable(gis = gis,
          itemID= itemID,
          lookupTablepath=lookupTablepath,
          file_name=file_name)

    # Part 2: Once lookup table is created and filled out, update on AGOL
    while True:
        answer = input("\n----> Please review the item's lookup table and modify the alias, description, and value type. Enter 'y' to proceed: ")
        if answer in ('yes', 'y'):
            print("Thank you! Proceeding to part 2 of uploading the changes on ArcGIS Online.\n")
            updateitemAGOL(lookupTablepath=lookupTablepath,
                           file_name=file_name,
                           itemID=itemID,
                           gis = gis)

            break
        else:
            print("Please enter y to proceed.")


if __name__ == "__main__":
    main()